# 챗봇을 똑똑하게 만들려면 데이터 학습이 필요하다.
# 챗봇 엔진은 입력되는 질문에 맞는 답변을 채택하기 위해
# 사전에 학습 데이터가 데이터베이스에 저장되어 있어야 한다.
# 챗봇의 학습 데이터를 데이터베이스에 저장하기 위해 학습 프로그램을 만드는 것도 좋지만
# 간편하게 정해진 엑셀 양식에 질문과 답변을 정리해서 데이터베이스에 저장할 것이다.

# 이미 학습된 데이터를 덤프 뜨거나 어떤 작업의 보고서를 만들 때 엑셀 형태로 저장하면
# 데이터를 해석할 때 엑셀 기능을 다양하게 활용할 수 있다.

# OpenPyXL 모듈은 파이썬에서 기본적으로 제공하지 않기 떄문에 모듈을 직접 설치해야 한다.
# 엑셀 파일을 읽고 쓸 수 있는 다른 라이브러리와 마찬가지로
# OpenPyXL 모듈에서도 엑셀 문서 구조에 대한 용어를 많이 사용한다.

# OpenPyXL 모듈 내 함수를 잘 사용하기 위해서는
# 엑셀 구조를 파악할 필요가 있다.

# 워크북은 엑셀 문서를 의미
# 워크북은 하나 이상의 워크시트를 가지고 있으며,
# 파일 명이 워크북의 이름이 된다.

# 워크 시트는 데이터를 입력할 수 있는 셀이
# 열과 행 형태로 구성되어 있는 엑셀 시트이다.

# 현재 활성화되어 있는 워크시트인 경우
# 액티브 시트라고 한다.

# 엑셀 파일의 특정 셀 내용 읽어오기
import openpyxl

# openpyxl 모듈 불러오기

wb = openpyxl.load_workbook('./sample.xlsx')
# 엑셀 파일을 워크북 인스턴스 객체로 저장하기

sheet = wb['Sheet1']
# 읽어온 워크북 객체에서
# 워크시트 이름을 인덱스를 사용해
# 워크시트 인스턴스 객체를 가져온다

print(sheet.max_column, sheet.max_row)
# 워크 시트에 저장되어 있는 열과 행의 크기 출력하기

print(sheet.cell(1, 1).value)
# cell() 함수를 이용해
# (1, 1) 위치의 워크시트 셀 데이터 내용을 가져와 출력
# cell() 함수에서 셀의 위치는 반드시 1부터 시작

print(sheet.cell(2, 1).value)
wb.close()  # 열었던 엑셀 파일을 닫기
# 사용을 완료한 리소스는 반드시 닫는다

"""
3 6
이름
Juyeon
"""

import openpyxl

wb = openpyxl.load_workbook('./sample.xlsx')
sheet = wb['Sheet1']

for row in sheet.iter_rows(min_row=2):
    for cell in row:        # 열 단위 데이터를 출력하기 위한 for 루프
        print(cell.value)   # 셀에 저장되어 있는 데이터 내용을 출력
    print('-' * 10)         # 하나의 행이 끝나는 부분을 표시하기 위해 구분선을 출력
wb.close()

"""
Juyeon
123
1234-1234
----------
Suyeon
45
1234-5678
----------
Hong
21
1234-4321
----------
Go
79
4321-7777
----------
Migz
98
4646-7979
----------
"""


import openpyxl

wb = openpyxl.load_workbook('./sample.xlsx')
sheet = wb['Sheet1']

cells = sheet['A2':'C3']    # 워크시트 객체에서 A2에서 C3 셀까지 슬라이싱해 지정한 범위만큼 셀 데이터 가져오기
for row in cells:
    for cell in row:        # 열 단위 데이터를 출력하기 위한 for 루프
        print(cell.value)   # 셀에 저장되어 있는 데이터 내용을 출력
    print('-' * 10)         # 하나의 행이 끝나는 부분을 표시하기 위해 구분선을 출력
wb.close()

"""
Juyeon
123
1234-1234
----------
Suyeon
45
1234-5678
----------
"""



# OpenPyXL로 엑셀 파일 쓰기
# 회원정보를 엑셀에 저장

import openpyxl

wb = openpyxl.Workbook()    # 워크북 인스턴스 객체 생성하기
sheet = wb.active           # 현재 활성화된 워크시트 객체 가져오기
sheet.title = '회원정보'      # 기본적으로 생성된 워크시트 제목은 'Sheet', 여기서는 '회원정보'

# 표 헤더 칼럼 저장
header_titles = ['이름', '전화번호']
for idx, title in enumerate(header_titles):
    sheet.cell(row=1, column=idx+1, value=title)

# 표 내용 저장
members=[('kei','010-1234-1234'), ('hong', '010-4321-1234')]
row_num=2
for r, member in enumerate(members):    # 회원정보 목록 탐색
    for c, v, in enumerate(member):     # 이름, 전화번호 칼럼 탐색
        sheet.cell(row=row_num+r, column=c+1, value=v)

wb.save('./members.xlsx')
wb.close()

"""
members.xlsx를 열어보자

이름  전화번호
kei  010-1234-1234
hong 010-4321-1234
"""